import pandas as pd
import re
import sys
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

PINK_FILL = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

def create_backup(file_path):
    """Create a timestamped backup of the given file."""
    file_path = Path(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
    backup_path = file_path.parent / backup_name
    shutil.copy2(file_path, backup_path)
    print(f"📁 Backup created: {backup_path}")
    return backup_path

def extract_component_keyword(comp_name):
    """Extract a meaningful keyword from the FML component name."""
    cleaned = comp_name.replace("CAT 6060", "").strip()
    return cleaned

def process_fml_file(fml_path, wb, ws, col_index):
    """Read FML file and update BARTRAC statuses."""
    print(f"Reading FML file: {fml_path}")
    df_fml_raw = pd.read_excel(fml_path, sheet_name="Sheet1", header=None, dtype=str)

    # Locate "Load Desc." row -> component names
    load_desc_idx = None
    for idx, row in df_fml_raw.iterrows():
        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if first_cell == "Load Desc.":
            load_desc_idx = idx
            break
    if load_desc_idx is None:
        raise ValueError("Could not find 'Load Desc.' row in FML file")
    components = [str(c).strip() for c in df_fml_raw.iloc[load_desc_idx, 1:] if pd.notna(c)]

    # Locate "Truck" row -> truck registrations
    truck_idx = None
    for idx, row in df_fml_raw.iterrows():
        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if first_cell == "Truck":
            truck_idx = idx
            break
    if truck_idx is None:
        raise ValueError("Could not find 'Truck' row in FML file")
    trucks = [str(t).strip() if pd.notna(t) else "" for t in df_fml_raw.iloc[truck_idx, 1:]]

    comp_to_truck = {comp: truck for comp, truck in zip(components, trucks) if truck}

    # Find the last date row (latest status)
    date_pattern = re.compile(r'^2026-\d{2}-\d{2}')
    data_rows = []
    for idx, row in df_fml_raw.iterrows():
        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if date_pattern.match(first_cell):
            data_rows.append(idx)
    if not data_rows:
        raise ValueError("No date rows found in FML file")
    last_data_idx = data_rows[-1]
    latest_statuses = [str(s).strip() if pd.notna(s) else "" for s in df_fml_raw.iloc[last_data_idx, 1:]]
    comp_to_status = {comp: status for comp, status in zip(components, latest_statuses) if comp}

    # Update BARTRAC rows
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=False):
        for cell in row:
            if cell.value and "CLIENT PO" in str(cell.value).upper():
                header_row_idx = cell.row
                break
        if header_row_idx is not None:
            break

    updated_count = 0
    for comp, truck_reg in comp_to_truck.items():
        keyword = extract_component_keyword(comp)
        status = comp_to_status.get(comp, "")
        if not status:
            print(f"⚠️ FML: No status found for component: {comp}")
            continue

        matching_rows = []
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=False):
            horse_cell = row[col_index["HORSE REG NO"] - 1]
            if horse_cell.value and str(horse_cell.value).strip() == truck_reg.strip():
                cargo_cell = row[col_index["CARGO DETAILS"] - 1]
                if cargo_cell.value and keyword.upper() in str(cargo_cell.value).upper():
                    matching_rows.append(row)
                else:
                    matching_rows.append(row)  # fallback

        if not matching_rows:
            print(f"❌ FML: No row found for truck '{truck_reg}' (component: {comp})")
            continue

        for row in matching_rows:
            status_cell = row[col_index["ACTUAL STATUS"] - 1]
            old_value = status_cell.value
            status_cell.value = status.upper()
            status_cell.fill = PINK_FILL
            updated_count += 1
            print(f"✅ FML: Updated row {row[0].row} for {comp} (truck {truck_reg}) → '{status.upper()}' (was: '{old_value}')")

    print(f"📊 FML total rows updated: {updated_count}")

def process_oriento_file(oriento_path, wb, ws, col_index):
    """Read ORIENTO tracking report (sheet BA3189) and update BARTRAC statuses."""
    print(f"Reading ORIENTO file: {oriento_path}")
    # Read raw without header
    df_raw = pd.read_excel(oriento_path, sheet_name="BA3189", header=None, dtype=str)

    # Find the row containing "TRUCK NUMBER" (header row)
    header_idx = None
    for idx, row in df_raw.iterrows():
        for cell in row.iloc[:20]:
            if pd.notna(cell) and "TRUCK NUMBER" in str(cell).upper():
                header_idx = idx
                break
        if header_idx is not None:
            break
    if header_idx is None:
        raise ValueError("Could not find header row in ORIENTO sheet BA3189")

    # Extract column indices for TRUCK NUMBER and CURRENT STATUS
    header_row = df_raw.iloc[header_idx]
    truck_col = None
    status_col = None
    for i, val in enumerate(header_row):
        if pd.notna(val) and "TRUCK NUMBER" in str(val).upper():
            truck_col = i
        elif pd.notna(val) and "CURRENT STATUS" in str(val).upper():
            status_col = i
    if truck_col is None or status_col is None:
        raise ValueError("Could not find 'TRUCK NUMBER' or 'CURRENT STATUS' columns in ORIENTO header")

    # Read data rows (after header)
    data_rows = df_raw.iloc[header_idx + 1:].copy()
    data_rows.columns = range(data_rows.shape[1])  # reset column indices

    # Build mapping truck -> status (using the latest row for each truck if duplicates)
    truck_to_status = {}
    for _, row in data_rows.iterrows():
        truck = str(row.iloc[truck_col]).strip() if pd.notna(row.iloc[truck_col]) else ""
        status = str(row.iloc[status_col]).strip() if pd.notna(row.iloc[status_col]) else ""
        if truck and status and truck != 'nan' and status != 'nan':
            truck_to_status[truck] = status

    print(f"Found {len(truck_to_status)} truck entries in ORIENTO report")

    # Find BARTRAC header row again (to know where data starts)
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=False):
        for cell in row:
            if cell.value and "CLIENT PO" in str(cell.value).upper():
                header_row_idx = cell.row
                break
        if header_row_idx is not None:
            break

    updated_count = 0
    for truck_reg, status in truck_to_status.items():
        # Find rows in BARTRAC that match this truck registration
        matching_rows = []
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=False):
            horse_cell = row[col_index["HORSE REG NO"] - 1]
            if horse_cell.value and str(horse_cell.value).strip() == truck_reg.strip():
                matching_rows.append(row)

        if not matching_rows:
            print(f"❌ ORIENTO: No row found for truck '{truck_reg}'")
            continue

        for row in matching_rows:
            status_cell = row[col_index["ACTUAL STATUS"] - 1]
            old_value = status_cell.value
            status_cell.value = status.upper()
            status_cell.fill = PINK_FILL
            updated_count += 1
            print(f"✅ ORIENTO: Updated row {row[0].row} (truck {truck_reg}) → '{status.upper()}' (was: '{old_value}')")

    print(f"📊 ORIENTO total rows updated: {updated_count}")

def main(fml_path, bartrac_path, oriento_path):
    # ------------------- 1. Create backup of BARTRAC file -------------------
    create_backup(bartrac_path)

    # ------------------- 2. Open BARTRAC workbook -------------------
    print(f"Opening BARTRAC file: {bartrac_path}")
    wb = load_workbook(bartrac_path)
    ws = wb["ENROUTE SITE"]

    # Find header row and column mapping
    header_row_idx = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=False):
        for cell in row:
            if cell.value and "CLIENT PO" in str(cell.value).upper():
                header_row_idx = cell.row
                header_row = row
                break
        if header_row_idx is not None:
            break
    if header_row_idx is None:
        raise ValueError("Could not find header row containing 'CLIENT PO' in ENROUTE SITE sheet")

    col_index = {}
    for col_idx, cell in enumerate(header_row, start=1):
        if cell.value:
            col_name = str(cell.value).strip()
            col_index[col_name] = col_idx

    required = ["CARGO DETAILS", "HORSE REG NO", "ACTUAL STATUS"]
    for col in required:
        if col not in col_index:
            raise ValueError(f"Column '{col}' not found in header row")

    # ------------------- 3. Process FML file -------------------
    process_fml_file(fml_path, wb, ws, col_index)

    # ------------------- 4. Process ORIENTO file -------------------
    process_oriento_file(oriento_path, wb, ws, col_index)

    # ------------------- 5. Save the modified workbook -------------------
    print(f"💾 Saving changes to original file: {bartrac_path}")
    wb.save(bartrac_path)

if __name__ == "__main__":
    default_fml = r"C:\Users\Jason\Projects\TRACKING-TASKS2\vendor-report\FML CAT 6060.xlsx"
    default_bartrac = "BARTRAC - KCC TRACKING AS OF 15-06-2026.xlsx"
    default_oriento = "ORIENTO TRACKING REPORT TO FML (DURBAN PORT TO KCC_.xlsx"

    fml_file = sys.argv[1] if len(sys.argv) > 1 else default_fml
    bartrac_file = sys.argv[2] if len(sys.argv) > 2 else default_bartrac
    oriento_file = sys.argv[3] if len(sys.argv) > 3 else default_oriento

    fml_path = Path(fml_file)
    bartrac_path = Path(bartrac_file)
    oriento_path = Path(oriento_file)

    main(fml_path, bartrac_path, oriento_path)