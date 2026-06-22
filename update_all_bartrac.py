# -----cab 220626 0820 ---
# ------1.run the directory vendor script directory_structure_vendor.py to get an output of files for
# -----update_all script to run---
# -----2.then check that the colour coding is done, re-sweep back to update any loos ends-
# -----3.copy the final files back into live---
# -----4.pray that this save time!!!---
import pandas as pd
import re
import sys
import shutil
import json
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

PINK_FILL = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

# ------------------- CONFIGURATION -------------------
BARTRAC_FOLDER = r"C:\Users\Jason\Projects\TRACKING-TASKS2\BARTRAC-TRACKING"
VENDOR_FOLDER = r"C:\Users\Jason\Projects\TRACKING-TASKS2\vendor-report"

HORSE_OVERRIDE_JSON = r"C:\Users\Jason\Projects\TRACKING-TASKS2\horse-registration.json"
SCRIPT_DIR = Path(__file__).resolve().parent
VENDOR_INDEX_JSON = SCRIPT_DIR / "directory_structure.json"

# ------------------------------------------------------------------

def create_backup(file_path):
    file_path = Path(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
    backup_path = file_path.parent / backup_name
    shutil.copy2(file_path, backup_path)
    print(f"📁 Backup created: {backup_path}")
    return backup_path

def load_vendor_index(index_path):
    index_path = Path(index_path)
    categories = {
        "ORIENTO_FILES": [],
        "FML_FILE": [],
        "NATRANS_FILE": [],
        "VANITO_FILE": []
    }

    if not index_path.exists():
        return categories

    try:
        with open(index_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"⚠️ Could not load vendor index {index_path}: {e}")
        return categories

    def traverse(node):
        if isinstance(node, dict):
            if node.get("type") == "file" and node.get("category") in categories and node.get("path"):
                categories[node["category"]].append(Path(node["path"]))
            for child in node.get("children", []):
                traverse(child)
        elif isinstance(node, list):
            for child in node:
                traverse(child)

    traverse(data)
    return categories

def get_sheet_name_for_bartrac(bartrac_path):
    filename = bartrac_path.name.lower()
    if "congo" in filename:
        return "CURRENT SHIPMENTS"
    elif "erg" in filename:
        return "CURRENT SHIPMENTS"
    elif "kamoa" in filename:
        return "ENROUTE SITE"
    elif "kcc" in filename:
        return "ENROUTE SITE"
    elif "mumi" in filename:
        return "ENROUTE SITE"
    else:
        return None

# ----- UPDATED: process_oriento_file now handles all sheets -----
def process_oriento_file(oriento_path):
    """Extract truck → status mapping from ALL sheets in the ORIENTO file.
    Returns {truck_rego: status_string}."""
    truck_to_status = {}
    oriento_path = Path(oriento_path)

    if not oriento_path.exists():
        print(f"⚠️ ORIENTO file not found: {oriento_path}")
        return truck_to_status

    try:
        xl = pd.ExcelFile(oriento_path)
    except Exception as e:
        print(f"⚠️ Could not read ORIENTO file {oriento_path}: {e}")
        return truck_to_status

    for sheet_name in xl.sheet_names:
        print(f"📖 Reading ORIENTO sheet: '{sheet_name}'")
        try:
            df_raw = pd.read_excel(oriento_path, sheet_name=sheet_name, header=None, dtype=str)
        except Exception as e:
            print(f"⚠️ Could not read sheet '{sheet_name}': {e}")
            continue

        # Find header row containing "TRUCK NUMBER" or similar
        header_idx = None
        header_keywords = ["TRUCK NUMBER", "TRUCK NO", "TRUCK", "HORSE", "HORSE REG", "HORSE REG NO"]
        for idx, row in df_raw.iterrows():
            for cell in row.iloc[:40]:
                if pd.notna(cell):
                    cell_upper = str(cell).upper()
                    if any(kw in cell_upper for kw in header_keywords):
                        header_idx = idx
                        break
            if header_idx is not None:
                break

        if header_idx is None:
            print(f"⚠️ Could not find header row in sheet '{sheet_name}'")
            continue

        header_row = df_raw.iloc[header_idx]
        truck_col = None
        location_col = None
        status_col = None

        for i, val in enumerate(header_row):
            if pd.notna(val):
                val_str = str(val).strip().upper()
                # Identify truck column
                if any(x in val_str for x in ["TRUCK NUMBER", "TRUCK NO", "TRUCK", "HORSE", "HORSE REG"]):
                    truck_col = i
                # Identify location column
                if any(x in val_str for x in ["CURRENT LOCATION", "LOCATION", "DESTINATION", "LOADING PLACE", "SITE"]):
                    location_col = i
                # Identify status column
                if any(x in val_str for x in ["CURRENT STATUS", "STATUS", "REMARKS", "COMMENT", "COMMENTS"]):
                    status_col = i

        if truck_col is None:
            print(f"⚠️ Could not find 'TRUCK'/'HORSE' column in sheet '{sheet_name}'")
            continue

        # Fallback if both location and status are missing
        if location_col is None and status_col is None:
            max_cols = df_raw.shape[1]
            cand_loc = truck_col + 1 if truck_col + 1 < max_cols else truck_col - 1
            cand_stat = truck_col + 2 if truck_col + 2 < max_cols else truck_col + 1 if truck_col + 1 < max_cols else truck_col - 1
            sample_row = df_raw.iloc[header_idx + 1] if header_idx + 1 < len(df_raw) else None
            used_fallback = False
            if sample_row is not None:
                if cand_loc is not None and pd.notna(sample_row.iloc[cand_loc]) and str(sample_row.iloc[cand_loc]).strip() not in ['', 'nan']:
                    location_col = cand_loc
                    used_fallback = True
                if cand_stat is not None and pd.notna(sample_row.iloc[cand_stat]) and str(sample_row.iloc[cand_stat]).strip() not in ['', 'nan']:
                    status_col = cand_stat
                    used_fallback = True
            if not used_fallback:
                print(f"⚠️ Could not find location or status columns in sheet '{sheet_name}', skipping.")
                continue

        # Process data rows
        data_rows = df_raw.iloc[header_idx + 1:].copy()
        data_rows.columns = range(data_rows.shape[1])

        for _, row in data_rows.iterrows():
            truck = str(row.iloc[truck_col]).strip() if pd.notna(row.iloc[truck_col]) else ""
            if not truck or truck == 'nan':
                continue

            location = str(row.iloc[location_col]).strip() if location_col is not None and pd.notna(row.iloc[location_col]) else ""
            status = str(row.iloc[status_col]).strip() if status_col is not None and pd.notna(row.iloc[status_col]) else ""

            if location and status:
                combined = f"{location} - {status}"
            elif location:
                combined = location
            elif status:
                combined = status
            else:
                continue

            truck_to_status[truck] = combined

    print(f"✅ ORIENTO: Found {len(truck_to_status)} truck entries")
    return truck_to_status

# ----- The rest of the functions remain unchanged -----
def process_fml_file(fml_path):
    """Extract truck → status mapping from FML file.
    Returns {truck_rego: status_string}."""
    truck_to_status = {}
    
    if not Path(fml_path).exists():
        print(f"⚠️ FML file not found: {fml_path}")
        return truck_to_status
    
    print(f"📖 Reading FML: {Path(fml_path).name}")
    try:
        df_fml_raw = pd.read_excel(fml_path, sheet_name="Sheet1", header=None, dtype=str)
    except Exception as e:
        print(f"❌ Could not read FML file: {e}")
        return truck_to_status

    load_desc_idx = None
    for idx, row in df_fml_raw.iterrows():
        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if first_cell == "Load Desc.":
            load_desc_idx = idx
            break
    
    if load_desc_idx is None:
        print(f"❌ Could not find 'Load Desc.' row in FML file")
        return truck_to_status
    
    components = [str(c).strip() for c in df_fml_raw.iloc[load_desc_idx, 1:] if pd.notna(c)]

    truck_idx = None
    for idx, row in df_fml_raw.iterrows():
        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if first_cell == "Truck":
            truck_idx = idx
            break
    
    if truck_idx is None:
        print(f"❌ Could not find 'Truck' row in FML file")
        return truck_to_status
    
    trucks = [str(t).strip() if pd.notna(t) else "" for t in df_fml_raw.iloc[truck_idx, 1:]]
    comp_to_truck = {comp: truck for comp, truck in zip(components, trucks) if truck}

    date_pattern = re.compile(r'^2026-\d{2}-\d{2}')
    data_rows = []
    for idx, row in df_fml_raw.iterrows():
        first_cell = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        if date_pattern.match(first_cell):
            data_rows.append(idx)
    
    if not data_rows:
        print(f"❌ No date rows found in FML file")
        return truck_to_status
    
    last_data_idx = data_rows[-1]
    latest_statuses = [str(s).strip() if pd.notna(s) else "" for s in df_fml_raw.iloc[last_data_idx, 1:]]
    comp_to_status = {comp: status for comp, status in zip(components, latest_statuses) if comp}

    for comp, truck_reg in comp_to_truck.items():
        status = comp_to_status.get(comp, "")
        if not status:
            continue
        truck_to_status[truck_reg] = status

    print(f"✅ FML: Found {len(truck_to_status)} truck entries")
    return truck_to_status

def process_natrans_file(natrans_path):
    """Extract truck → status mapping from NATRANS file.
    Returns {truck_rego: status_string}."""
    truck_to_status = {}
    natrans_path = Path(natrans_path)
    
    if not natrans_path.exists():
        print(f"⚠️ NATRANS file not found: {natrans_path}")
        return truck_to_status

    print(f"📖 Reading NATRANS: {natrans_path.name}")
    try:
        df_raw = pd.read_excel(natrans_path, sheet_name="NATRANS", header=None, dtype=str)
    except ValueError:
        print(f"⚠️ Sheet 'NATRANS' not found in {natrans_path.name}. Skipping.")
        return truck_to_status
    except Exception as e:
        print(f"❌ Could not read NATRANS file: {e}")
        return truck_to_status

    header_idx = None
    for idx, row in df_raw.iterrows():
        for cell in row.iloc[:10]:
            if pd.notna(cell) and str(cell).strip().upper() in ["LOAD", "DISP", "BB SA", "BB ZIM"]:
                header_idx = idx
                break
        if header_idx is not None:
            break
    
    if header_idx is None:
        print(f"❌ Could not find header row in NATRANS sheet")
        return truck_to_status

    header_row = df_raw.iloc[header_idx]
    col_names = {}
    for i, val in enumerate(header_row):
        if pd.notna(val):
            col_names[str(val).strip()] = i

    horse_col = None
    for possible in ["HORSE", "HORSE REG NO", "HORSE REG"]:
        if possible in col_names:
            horse_col = col_names[possible]
            break
    
    if horse_col is None:
        print(f"❌ Could not find HORSE column in NATRANS file")
        return truck_to_status

    milestone_columns = []
    for col in ["LOAD", "DISP", "BB SA", "BB ZIM", "CHI ZIM", "CHI ZAM", 
                "KASUM ZAM", "KASUM DRC", "WHISKEY", "SITE", "OFFLOAD", 
                "LUFUA ARR", "LUFUA DISP"]:
        if col in col_names:
            milestone_columns.append(col)

    data_rows = df_raw.iloc[header_idx + 1:].copy()
    data_rows.columns = range(data_rows.shape[1])

    for _, row in data_rows.iterrows():
        truck = str(row.iloc[horse_col]).strip() if pd.notna(row.iloc[horse_col]) else ""
        if not truck or truck == 'nan':
            continue

        last_status = None
        for milestone in milestone_columns:
            col_idx = col_names[milestone]
            cell_value = row.iloc[col_idx]
            if pd.notna(cell_value) and str(cell_value).strip() not in ['', 'nan']:
                last_status = milestone

        if "COMMENTS" in col_names:
            comments_idx = col_names["COMMENTS"]
            comments_val = row.iloc[comments_idx]
            if pd.notna(comments_val) and str(comments_val).strip() not in ['', 'nan']:
                truck_to_status[truck] = str(comments_val).strip()
            elif last_status:
                truck_to_status[truck] = last_status
        else:
            if last_status:
                truck_to_status[truck] = last_status

    print(f"✅ NATRANS: Found {len(truck_to_status)} truck entries")
    return truck_to_status

def process_vanito_file(vanito_path):
    """Extract truck → status mapping from Vanito file.
    Returns {truck_rego: status_string}."""
    truck_to_status = {}
    vanito_path = Path(vanito_path)
    
    if not vanito_path.exists():
        print(f"⚠️ Vanito file not found: {vanito_path}")
        return truck_to_status

    print(f"📖 Reading Vanito: {vanito_path.name}")

    try:
        xl = pd.ExcelFile(vanito_path)
        sheets = xl.sheet_names
    except Exception as e:
        print(f"❌ Could not read sheets from Vanito file: {e}")
        return truck_to_status

    date_pattern = re.compile(r'^(\d{1,2})\s+([A-Za-z]{3,})\s+(\d{4})$')
    sheet_date_map = {}
    for sheet in sheets:
        m = date_pattern.match(sheet.strip())
        if m:
            day, month_str, year = m.groups()
            try:
                month_num = datetime.strptime(month_str.title(), "%b").month
                dt = datetime(int(year), month_num, int(day))
                sheet_date_map[dt] = sheet
            except ValueError:
                continue

    if not sheet_date_map:
        print("⚠️ No sheets with date format found in Vanito file.")
        return truck_to_status

    latest_date = max(sheet_date_map.keys())
    latest_sheet = sheet_date_map[latest_date]
    print(f"Using latest sheet: '{latest_sheet}' (date: {latest_date.strftime('%Y-%m-%d')})")

    try:
        df_raw = pd.read_excel(vanito_path, sheet_name=latest_sheet, header=None, dtype=str)
    except Exception as e:
        print(f"❌ Could not read sheet '{latest_sheet}': {e}")
        return truck_to_status

    header_idx = None
    for idx, row in df_raw.iterrows():
        for cell in row.iloc[:20]:
            if pd.notna(cell) and "HORSE REG" in str(cell).upper():
                header_idx = idx
                break
        if header_idx is not None:
            break
    
    if header_idx is None:
        print(f"⚠️ Could not find header row with 'HORSE REG' in sheet '{latest_sheet}'")
        return truck_to_status

    header_row = df_raw.iloc[header_idx]
    horse_col = None
    location_col = None
    for i, val in enumerate(header_row):
        val_str = str(val).strip().upper()
        if "HORSE REG" in val_str:
            horse_col = i
        elif "CURRENT LOCATION" in val_str:
            location_col = i
        if location_col is None and "CURRENT STATUS" in val_str:
            location_col = i
        if location_col is None and "LOCATION" in val_str:
            location_col = i

    if horse_col is None:
        print(f"⚠️ Could not find 'HORSE REG' column in sheet '{latest_sheet}'")
        return truck_to_status
    
    if location_col is None:
        print(f"⚠️ Could not find 'CURRENT LOCATION' (or equivalent) column in sheet '{latest_sheet}'")
        return truck_to_status

    data_rows = df_raw.iloc[header_idx + 1:].copy()
    data_rows.columns = range(data_rows.shape[1])

    for _, row in data_rows.iterrows():
        truck = str(row.iloc[horse_col]).strip() if pd.notna(row.iloc[horse_col]) else ""
        location = str(row.iloc[location_col]).strip() if pd.notna(row.iloc[location_col]) else ""
        if truck and location and truck != 'nan' and location != 'nan':
            truck_to_status[truck] = location

    print(f"✅ Vanito: Found {len(truck_to_status)} truck entries")
    return truck_to_status

def load_horse_overrides(json_path):
    """Load manual horse status overrides from JSON file."""
    overrides = {}
    if not Path(json_path).exists():
        print(f"ℹ️ Horse registration file not found: {json_path}")
        return overrides
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"❌ Invalid JSON in {json_path}: {e}")
        return overrides

    for section, entries in data.items():
        if isinstance(entries, list):
            for entry in entries:
                if isinstance(entry, dict) and "code" in entry:
                    code = entry.get("code")
                    status = entry.get("status")
                    colour = entry.get("colour")
                    if code and status:
                        overrides[code] = {"status": status, "colour": colour}
    
    print(f"ℹ️ Loaded {len(overrides)} manual overrides from JSON")
    return overrides

def apply_status_map_to_bartrac(bartrac_path, status_map, overrides, dry_run=False):
    """Apply master status map to a single BARTRAC file.
    
    Args:
        bartrac_path: Path to BARTRAC file
        status_map: {truck_rego: status_string} dict from vendors
        overrides: {truck_rego: {status, colour}} dict from JSON
        dry_run: If True, print updates without saving

    Returns:
        Tuple[int, int] if processed successfully: (vendor_updates, override_updates)
        None if file could not be processed.
    """
    bartrac_path = Path(bartrac_path)
    print(f"\n{'='*60}")
    print(f"Processing: {bartrac_path.name}")
    print(f"{'='*60}")

    # Determine sheet name
    sheet_name = get_sheet_name_for_bartrac(bartrac_path)
    if sheet_name is None:
        try:
            wb = load_workbook(bartrac_path)
            found = None
            for s in wb.sheetnames:
                if s.strip() == "ENROUTE SITE":
                    found = s
                    break
            if found is None:
                for s in wb.sheetnames:
                    if s.strip() == "CURRENT SHIPMENTS":
                        found = s
                        break
            if found is not None:
                sheet_name = found
            else:
                print(f"❌ No known sheet found in {bartrac_path.name}")
                return None
        except Exception as e:
            print(f"❌ Cannot open {bartrac_path.name}: {e}")
            return None
    else:
        try:
            wb = load_workbook(bartrac_path)
            if sheet_name not in wb.sheetnames:
                for s in wb.sheetnames:
                    if s.strip() == sheet_name.strip():
                        sheet_name = s
                        break
                if sheet_name not in wb.sheetnames:
                    print(f"❌ Sheet '{sheet_name}' not found in {bartrac_path.name}")
                    return None
        except Exception as e:
            print(f"❌ Cannot open {bartrac_path.name}: {e}")
            return None

    print(f"Using sheet: '{sheet_name}'")

    if dry_run:
        print(f"DRY-RUN: Skipping backup for {bartrac_path.name}")
    else:
        create_backup(bartrac_path)

    wb = load_workbook(bartrac_path)
    ws = wb[sheet_name]

    # Find header row
    header_row_idx = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=False):
        for cell in row:
            if cell.value and "CLIENT PO" in str(cell.value).upper():
                header_row_idx = cell.row
                header_row = row
                break
        if header_row_idx is not None:
            break
    
    if header_row_idx is None:
        print(f"❌ Could not find header row containing 'CLIENT PO'")
        return None

    # Build column index
    col_index = {}
    for col_idx, cell in enumerate(header_row, start=1):
        if cell.value:
            col_name = str(cell.value).strip()
            col_index[col_name] = col_idx

    required = ["CARGO DETAILS", "HORSE REG NO", "ACTUAL STATUS"]
    for col in required:
        if col not in col_index:
            print(f"❌ Column '{col}' not found in header row")
            return None

    # Apply status map (vendor data)
    updated_count = 0
    for horse_reg, status in status_map.items():
        matching_rows = []
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=False):
            horse_cell = row[col_index["HORSE REG NO"] - 1]
            if horse_cell.value and str(horse_cell.value).strip() == horse_reg.strip():
                matching_rows.append(row)

        if not matching_rows:
            continue  # Silently skip if truck not found (may be in other BARTRAC files)

        for row in matching_rows:
            status_cell = row[col_index["ACTUAL STATUS"] - 1]
            old_value = status_cell.value
            new_value = status.upper()
            
            if dry_run:
                print(f"DRY-RUN: Would update row {row[0].row} ({horse_reg}) → '{new_value}' (was: '{old_value}')")
            else:
                status_cell.value = new_value
                status_cell.fill = PINK_FILL
                print(f"✅ Updated row {row[0].row} ({horse_reg}) → '{new_value}'")
            updated_count += 1

    print(f"📊 Vendor data: {updated_count} rows updated")

    # Apply overrides (manual entries - highest priority)
    override_count = 0
    for horse_reg in overrides:
        matching_rows = []
        for row in ws.iter_rows(min_row=header_row_idx+1, values_only=False):
            horse_cell = row[col_index["HORSE REG NO"] - 1]
            if horse_cell.value and str(horse_cell.value).strip() == horse_reg.strip():
                matching_rows.append(row)

        if not matching_rows:
            continue

        override = overrides[horse_reg]
        for row in matching_rows:
            status_cell = row[col_index["ACTUAL STATUS"] - 1]
            old_value = status_cell.value
            new_value = override["status"].upper()
            
            if dry_run:
                print(f"DRY-RUN OVERRIDE: Row {row[0].row} ({horse_reg}) → '{new_value}' (was '{old_value}')")
            else:
                status_cell.value = new_value
                if override.get("colour") == "pink":
                    status_cell.fill = PINK_FILL
                print(f"🎨 OVERRIDE: Row {row[0].row} ({horse_reg}) → '{new_value}'")
            override_count += 1

    print(f"📌 Manual overrides: {override_count} rows updated")

    if dry_run:
        print(f"DRY-RUN: Changes not saved to {bartrac_path}")
    else:
        print(f"💾 Saving changes to: {bartrac_path}")
        wb.save(bartrac_path)

    return updated_count, override_count

def main():
    parser = argparse.ArgumentParser(description='Update BARTRAC files from vendor reports')
    parser.add_argument('--dry-run', action='store_true', help='Print intended updates without saving changes')
    args = parser.parse_args()
    dry_run = args.dry_run

    vendor_index = load_vendor_index(VENDOR_INDEX_JSON)
    has_vendor_index = any(vendor_index.values())
    if has_vendor_index:
        print(f"ℹ️ Loaded vendor file list from {VENDOR_INDEX_JSON}")

    # Vendor folder for discovery fallback
    vendor_folder = Path(VENDOR_FOLDER)

    # ========== 1. FML files (as a list) ==========
    fml_paths = []
    if has_vendor_index and vendor_index["FML_FILE"]:
        fml_paths = [p for p in vendor_index["FML_FILE"] if p.exists()]
        print(f"ℹ️ Using FML files from {VENDOR_INDEX_JSON} ({len(fml_paths)} entries)")

    if not fml_paths and vendor_folder.exists():
        fml_candidates = [p for p in vendor_folder.glob('*.xlsx') 
                          if 'FML' in p.name.upper() and not p.name.startswith('~$')]
        if fml_candidates:
            # Optional: filter to only those containing '6060' or 'CAT'
            # cat_candidates = [p for p in fml_candidates if '6060' in p.name.upper() or 'CAT' in p.name.upper()]
            # fml_paths = cat_candidates if cat_candidates else fml_candidates
            fml_paths = fml_candidates

    if not fml_paths:
        print("ERROR: No FML files found.")
        if not dry_run:
            sys.exit(1)

    # ========== 2. ORIENTO files (list) ==========
    oriento_paths = []
    if has_vendor_index and vendor_index["ORIENTO_FILES"]:
        oriento_paths = [p for p in vendor_index["ORIENTO_FILES"] if p.exists()]
        print(f"ℹ️ Using ORIENTO files from {VENDOR_INDEX_JSON} ({len(oriento_paths)} entries)")
    if not oriento_paths and vendor_folder.exists():
        discovered = sorted([p for p in vendor_folder.glob('*.xlsx') if 'ORIENTO' in p.name.upper() or 'CATERPILLAR' in p.name.upper()])
        for p in discovered:
            if p not in oriento_paths:
                oriento_paths.append(p)

    # ========== 3. NATRANS files (list) ==========
    natrans_paths = []
    if has_vendor_index and vendor_index["NATRANS_FILE"]:
        natrans_paths = [p for p in vendor_index["NATRANS_FILE"] if p.exists()]
        print(f"ℹ️ Using NATRANS files from {VENDOR_INDEX_JSON} ({len(natrans_paths)} entries)")
    if not natrans_paths and vendor_folder.exists():
        natrans_candidates = [p for p in vendor_folder.glob('*.xlsx') if 'FML' in p.name.upper() and ('KOLWEZI' in p.name.upper() or 'SAKANIA' in p.name.upper() or 'NATRANS' in p.name.upper())]
        for p in natrans_candidates:
            if p not in natrans_paths:
                natrans_paths.append(p)

    # ========== 4. VANITO file (single) ==========
    vanito_path = None
    if has_vendor_index and vendor_index["VANITO_FILE"]:
        for p in vendor_index["VANITO_FILE"]:
            if p.exists():
                vanito_path = p
                break
    if vanito_path is None and vendor_folder.exists():
        vanito_candidates = [p for p in vendor_folder.glob('*.xlsx') if 'VANITO' in p.name.upper() or 'FML DBN TO DRC' in p.name.upper()]
        if vanito_candidates:
            vanito_path = vanito_candidates[0]

    # ========== 5. BARTRAC folder ==========
    bartrac_folder = Path(BARTRAC_FOLDER)
    horse_json = Path(HORSE_OVERRIDE_JSON)

    if not bartrac_folder.exists():
        print(f"ERROR: BARTRAC folder does not exist: {bartrac_folder}")
        sys.exit(1)

    bartrac_files = list(bartrac_folder.glob("*.xlsx"))
    bartrac_files = [f for f in bartrac_files if not f.name.startswith("~$")]
    if not bartrac_files:
        print(f"No .xlsx files found in {bartrac_folder}")
        sys.exit(0)

    print(f"\n{'='*60}")
    print(f"PHASE 1: Building master status map from vendors")
    print(f"{'='*60}\n")

    # ========== PHASE 1: Build master status map ==========
    master_status = {}

    # 1) FML files (all)
    for fml_path in fml_paths:
        status_from_fml = process_fml_file(fml_path)
        master_status.update(status_from_fml)

    # 2) ORIENTO files (updated to read all sheets)
    for oriento_path in oriento_paths:
        status_from_oriento = process_oriento_file(oriento_path)
        master_status.update(status_from_oriento)

    # 3) NATRANS files
    for natrans_path in natrans_paths:
        status_from_natrans = process_natrans_file(natrans_path)
        master_status.update(status_from_natrans)

    # 4) VANITO file (if found)
    if vanito_path is not None and Path(vanito_path).exists():
        status_from_vanito = process_vanito_file(vanito_path)
        master_status.update(status_from_vanito)
    else:
        print("⚠️ Vanito file not found, skipping.")
        status_from_vanito = {}

    print(f"\n📊 Master status map contains {len(master_status)} unique truck entries\n")

    # ========== PHASE 2: Load overrides ==========
    overrides = load_horse_overrides(horse_json)

    # ========== PHASE 3: Apply to BARTRAC files ==========
    print(f"\n{'='*60}")
    print(f"PHASE 2: Applying status map to BARTRAC files")
    print(f"{'='*60}")

    print(f"\nFound {len(bartrac_files)} BARTRAC file(s) to process.\n")
    summary_rows = []
    for bartrac_file in bartrac_files:
        try:
            result = apply_status_map_to_bartrac(bartrac_file, master_status, overrides, dry_run=dry_run)
            if result is not None:
                vendor_updates, override_updates = result
                summary_rows.append((bartrac_file.name, vendor_updates, override_updates))
            else:
                summary_rows.append((bartrac_file.name, None, None))
        except Exception as e:
            print(f"❌ Failed to process {bartrac_file.name}: {e}")
            summary_rows.append((bartrac_file.name, None, None))

    print(f"\n✅ All BARTRAC files processed.")
    print("\nSummary of row changes per file:")
    for filename, vendor_updates, override_updates in summary_rows:
        if vendor_updates is None:
            print(f"- {filename}: failed to process or no known sheet/header")
        else:
            print(f"- {filename}: vendor updates={vendor_updates}, manual overrides={override_updates}")

if __name__ == "__main__":
    main()